from __future__ import annotations

import argparse
import hashlib
import json
import logging
import random
import re
import time
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Set
from urllib.parse import urljoin, urlparse, parse_qs, urlencode

import httpx
import portalocker
from openpyxl import Workbook, load_workbook
from playwright.sync_api import TimeoutError as PlaywrightTimeoutError
from playwright.sync_api import sync_playwright

from app.config import Settings
from app.scraping.discover_api import (
    decode_response_json,
    extract_products_from_payload,
    looks_like_product_endpoint,
)

log = logging.getLogger(__name__)

TARGET_URL = "https://www.bazarey.store/en/product"
PRODUCT_HEADERS = [
    "product_id",
    "name",
    "url",
    "category",
    "price",
    "regular_price",
    "unit",
    "stock_qty",
    "is_active",
    "image_url",
    "updated_at",
]

# Regex to extract MongoDB-style hex ID from product URLs
_PRODUCT_URL_ID_RE = re.compile(r"/product/([a-f0-9]{24})(?:\?|$|#)")


def utc_now_iso() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat()


def parse_bool(text: str) -> bool:
    return str(text).strip().lower() in {"1", "true", "yes", "y"}


def to_price(value: Any) -> float | None:
    if value is None or value == "":
        return None
    try:
        if isinstance(value, str):
            cleaned = re.sub(r"[^\d.,-]", "", value).replace(",", "")
            if not cleaned:
                return None
            return float(cleaned)
        return float(value)
    except Exception:
        return None


def stable_product_id(url: str) -> str:
    """Extract MongoDB ObjectID from URL, or fall back to SHA1 hash."""
    m = _PRODUCT_URL_ID_RE.search(url)
    if m:
        return m.group(1)
    return hashlib.sha1(url.encode("utf-8")).hexdigest()[:14]


def ensure_products_file(path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if path.exists():
        return
    wb = Workbook()
    ws = wb.active
    ws.title = "products"
    ws.append(PRODUCT_HEADERS)
    wb.save(path)


def load_existing(path: Path) -> Dict[str, Dict[str, Any]]:
    ensure_products_file(path)
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    out: Dict[str, Dict[str, Any]] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        item = dict(zip(headers, row))
        pid = str(item.get("product_id") or "").strip()
        url = str(item.get("url") or "").strip()
        key = pid or url
        if key:
            out[key] = item
    wb.close()
    return out


def write_products(path: Path, rows: List[Dict[str, Any]]) -> None:
    ensure_products_file(path)
    lock_path = path.with_suffix(path.suffix + ".lock")
    with portalocker.Lock(lock_path, timeout=10):
        wb = Workbook()
        ws = wb.active
        ws.title = "products"
        ws.append(PRODUCT_HEADERS)
        for row in rows:
            ws.append([row.get(h, "") for h in PRODUCT_HEADERS])
        wb.save(path)


def normalize_product(raw: Dict[str, Any], base_url: str) -> Dict[str, Any] | None:
    name = str(raw.get("name") or "").strip()
    url = str(raw.get("url") or "").strip()
    if url and not url.startswith("http"):
        url = urljoin(base_url, url)
    price = to_price(raw.get("price"))
    if not name or price is None:
        return None

    product_id = str(raw.get("product_id") or "").strip() or stable_product_id(url or name)
    return {
        "product_id": product_id,
        "name": name,
        "url": url,
        "category": str(raw.get("category") or "").strip(),
        "price": price,
        "regular_price": to_price(raw.get("regular_price")),
        "unit": str(raw.get("unit") or "").strip(),
        "stock_qty": raw.get("stock_qty") or "",
        "is_active": bool(raw.get("is_active", True)),
        "image_url": str(raw.get("image_url") or "").strip(),
        "updated_at": utc_now_iso(),
    }


def try_goto(page: Any, url: str, retries: int = 3) -> None:
    last_err = None
    for _ in range(retries):
        try:
            page.goto(url, wait_until="networkidle", timeout=60_000)
            return
        except PlaywrightTimeoutError as exc:
            last_err = exc
            time.sleep(1.2)
    if last_err:
        raise last_err


def collect_dom_listing(page: Any, base_url: str) -> List[Dict[str, Any]]:
    """Extract product cards from the current page DOM.

    The Bazarey product listing cards follow this structure:
      - <a> tag with href containing /product/<ObjectID>
      - Parent card with product name, price (৳NNN), weight/unit text
      - Optional <img> element
    """
    cards = page.evaluate(
        """
        () => {
          // Match product detail links (24-char hex IDs)
          const anchors = [...document.querySelectorAll('a[href*="/product/"]')];
          const out = [];
          const seen = new Set();
          for (const a of anchors) {
            const href = a.getAttribute('href') || '';
            // Only pick links that point to a specific product (24-char hex ID)
            const idMatch = href.match(/\\/product\\/([a-f0-9]{24})/);
            if (!idMatch) continue;
            const productId = idMatch[1];
            if (seen.has(productId)) continue;
            seen.add(productId);

            const card = a.closest('article,div,li,section') || a.parentElement;
            const text = card ? card.innerText : a.innerText;
            const imgEl = card ? card.querySelector('img') : null;

            out.push({
              product_id: productId,
              name: (a.getAttribute('title') || a.innerText || '').trim(),
              url: href,
              text: (text || '').trim(),
              image_url: imgEl ? (imgEl.src || imgEl.getAttribute('data-src') || '') : ''
            });
          }
          return out;
        }
        """
    )

    products: List[Dict[str, Any]] = []
    seen: Set[str] = set()
    for card in cards:
        pid = str(card.get("product_id") or "").strip()
        url = str(card.get("url") or "").strip()
        if not url:
            continue
        full_url = urljoin(base_url, url)
        key = pid or full_url
        if key in seen:
            continue
        seen.add(key)

        text_blob = str(card.get("text") or "")

        # Extract price: look for ৳ symbol first, then bare numbers
        price_val = None
        price_match = re.search(r"৳\s*([\d,]+(?:\.\d+)?)", text_blob)
        if price_match:
            price_val = float(price_match.group(1).replace(",", ""))
        else:
            price_match = re.search(r"([\d,]+(?:\.\d+)?)", text_blob)
            if price_match:
                price_val = float(price_match.group(1).replace(",", ""))

        # Extract weight/unit from card text
        unit = ""
        unit_match = re.search(
            r"(?:Weight/Unit[:\s]*)?(\d+(?:\.\d+)?\s*(?:kg|g|gm|l|ml|pcs|pc|piece|liter|litre))",
            text_blob, flags=re.IGNORECASE,
        )
        if unit_match:
            unit = unit_match.group(1).strip()

        # Clean up name: remove duplicate text, price/unit noise
        raw_name = str(card.get("name") or "").strip()
        # The card text often duplicates the name; use the first line as primary
        if raw_name:
            # Remove trailing 'View' button text
            raw_name = re.sub(r"\s*View\s*$", "", raw_name, flags=re.IGNORECASE).strip()
            # If name is duplicated (e.g. "Product Name Product Name Weight...")
            # take just the first occurrence
            lines = [l.strip() for l in raw_name.split("\n") if l.strip()]
            if lines:
                raw_name = lines[0]

        products.append(
            {
                "product_id": pid or stable_product_id(full_url),
                "name": raw_name or f"Product {len(products) + 1}",
                "url": full_url,
                "price": price_val,
                "unit": unit,
                "image_url": str(card.get("image_url") or "").strip(),
            }
        )
    return products


def _get_total_pages(page: Any) -> int:
    """Detect total page count from pagination links."""
    try:
        max_page = page.evaluate(
            """
            () => {
              const links = [...document.querySelectorAll('a[href*="page="], button, nav a, [class*="pagination"] a, [class*="Pagination"] a')];
              let maxPage = 1;
              for (const el of links) {
                const text = (el.textContent || '').trim();
                const num = parseInt(text, 10);
                if (!isNaN(num) && num > maxPage && num < 200) {
                  maxPage = num;
                }
                // Also check href for page= parameter
                const href = el.getAttribute('href') || '';
                const match = href.match(/[?&]page=(\\d+)/);
                if (match) {
                  const pnum = parseInt(match[1], 10);
                  if (pnum > maxPage && pnum < 200) maxPage = pnum;
                }
              }
              return maxPage;
            }
            """
        )
        return max(1, int(max_page))
    except Exception:
        return 1


def _click_next_page(page: Any, target_page: int) -> bool:
    """Click a pagination button for the target page number.

    Tries multiple strategies:
    1. Click a visible link/button whose text is exactly the target page number.
    2. Click a 'Next' / '>' / '→' button.

    Returns True if page content appeared to change.
    """
    try:
        # Collect current product IDs so we can detect content change
        old_ids = set(page.evaluate(
            """() => [...document.querySelectorAll('a[href*="/product/"]')]
                      .map(a => a.href)
                      .filter(h => /\\/product\\/[a-f0-9]{24}/.test(h))"""
        ))

        # Strategy 1: click exact page number link
        clicked = page.evaluate(
            f"""(targetPage) => {{
              const links = [...document.querySelectorAll('a, button')];
              for (const el of links) {{
                const text = (el.textContent || '').trim();
                if (text === String(targetPage)) {{
                  el.scrollIntoView();
                  el.click();
                  return true;
                }}
              }}
              return false;
            }}""",
            target_page,
        )

        if not clicked:
            # Strategy 2: click Next / > / → button
            clicked = page.evaluate(
                """() => {
                  const links = [...document.querySelectorAll('a, button')];
                  for (const el of links) {
                    const text = (el.textContent || '').trim().toLowerCase();
                    const ariaLabel = (el.getAttribute('aria-label') || '').toLowerCase();
                    if (text === '>' || text === '→' || text === 'next' ||
                        text === '›' || ariaLabel.includes('next')) {
                      el.scrollIntoView();
                      el.click();
                      return true;
                    }
                  }
                  return false;
                }"""
            )

        if not clicked:
            return False

        # Wait for content to update by polling for new product links
        for _ in range(20):
            time.sleep(0.5)
            new_ids = set(page.evaluate(
                """() => [...document.querySelectorAll('a[href*="/product/"]')]
                          .map(a => a.href)
                          .filter(h => /\\/product\\/[a-f0-9]{24}/.test(h))"""
            ))
            if new_ids and new_ids != old_ids:
                return True

        # Even if content didn't change, let subsequent collection handle it
        return True
    except Exception as exc:
        log.warning("Failed to click page %d: %s", target_page, exc)
        return False


def scrape_product_detail(page: Any, product: Dict[str, Any], sleep_min: float = 0.25, sleep_max: float = 0.9) -> Dict[str, Any] | None:
    if not product.get("url"):
        return normalize_product(product, TARGET_URL)

    try:
        try_goto(page, product["url"], retries=2)
    except Exception:
        return normalize_product(product, TARGET_URL)

    name = page.locator("h1").first.text_content(timeout=2_000) if page.locator("h1").count() else product.get("name")
    body = page.locator("body").inner_text(timeout=2_000)
    price = to_price(product.get("price"))
    if price is None:
        match = re.search(r"(?:৳|Tk|BDT)\s*([\d,]+(?:\.\d+)?)", body, flags=re.IGNORECASE)
        if match:
            price = float(match.group(1).replace(",", ""))

    # Extract unit/weight
    unit = product.get("unit") or ""
    if not unit:
        match_unit = re.search(
            r"(?:Weight/Unit[:\s]*)?(\d+(?:\.\d+)?\s*(?:kg|g|gm|l|ml|pcs|pc|piece|liter|litre))",
            body, flags=re.IGNORECASE,
        )
        if match_unit:
            unit = match_unit.group(1)

    # Extract category from breadcrumb or body text
    category = product.get("category", "")
    if not category:
        cat_match = re.search(r"Categor(?:y|ies)\s*[:\s]*([A-Za-z &,\u0980-\u09FF]+)", body)
        if cat_match:
            category = cat_match.group(1).strip()

    # Extract stock status
    stock_qty = product.get("stock_qty", "")
    if not stock_qty:
        if re.search(r"\bIn\s*stock\b", body, flags=re.IGNORECASE):
            stock_qty = "In Stock"
        elif re.search(r"\bOut\s*of\s*stock\b", body, flags=re.IGNORECASE):
            stock_qty = "Out of Stock"

    image_url = product.get("image_url") or ""
    if not image_url and page.locator("img").count():
        img_src = page.locator("img").first.get_attribute("src")
        if img_src:
            image_url = urljoin(TARGET_URL, img_src)

    normalized = normalize_product(
        {
            "product_id": product.get("product_id"),
            "name": name or product.get("name"),
            "url": product.get("url"),
            "category": category,
            "price": price,
            "regular_price": product.get("regular_price"),
            "unit": unit,
            "stock_qty": stock_qty,
            "is_active": True,
            "image_url": image_url,
        },
        TARGET_URL,
    )
    time.sleep(random.uniform(sleep_min, sleep_max))
    return normalized


def scrape_from_discovered_endpoints(endpoints: List[str]) -> List[Dict[str, Any]]:
    out: List[Dict[str, Any]] = []
    paginated: List[Dict[str, Any]] = []
    if not endpoints:
        return out

    successful_endpoints: List[str] = []
    with httpx.Client(timeout=20) as client:
        for endpoint in endpoints:
            try:
                r = client.get(endpoint)
                r.raise_for_status()
                payload = r.json()
            except Exception:
                log.warning("Failed to fetch endpoint: %s", endpoint)
                continue
            products = extract_products_from_payload(payload, source_url=endpoint)
            out.extend(products)
            if products:
                successful_endpoints.append(endpoint)

    # Try paginating only endpoints that succeeded
    for endpoint in successful_endpoints:
        parsed = urlparse(endpoint)
        qs = parse_qs(parsed.query)
        if "page" not in qs and "limit" not in qs:
            continue
        with httpx.Client(timeout=20) as client:
            for page_num in range(2, 100):
                qs_copy = dict(qs)
                qs_copy["page"] = [str(page_num)]
                new_url = parsed._replace(query=urlencode(qs_copy, doseq=True)).geturl()
                try:
                    r = client.get(new_url)
                    r.raise_for_status()
                    payload = r.json()
                    products = extract_products_from_payload(payload, source_url=new_url)
                    if not products:
                        break
                    paginated.extend(products)
                except Exception:
                    break
    out.extend(paginated)
    return out


def run_scraper(headless: bool, limit: int, output: Path, slowmo: int, api_endpoints_path: Path) -> int:
    existing = load_existing(output)
    discovered_products: List[Dict[str, Any]] = []
    endpoint_urls: Set[str] = set()
    all_listing_products: List[Dict[str, Any]] = []

    log.info("Starting scraper — headless=%s, limit=%d", headless, limit)

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=headless, slow_mo=slowmo)
        context = browser.new_context(
            user_agent="Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                       "AppleWebKit/537.36 (KHTML, like Gecko) "
                       "Chrome/122.0.0.0 Safari/537.36"
        )
        page = context.new_page()

        def on_response(resp: Any) -> None:
            try:
                ctype = (resp.headers or {}).get("content-type", "")
                if "application/json" not in ctype and not looks_like_product_endpoint(resp.url):
                    return
                body = resp.text()
                payload = decode_response_json(body)
                products = extract_products_from_payload(payload, source_url=resp.url)
                if products:
                    endpoint_urls.add(resp.url)
                    discovered_products.extend(products)
            except Exception:
                return

        page.on("response", on_response)

        # Load first page and detect total page count
        log.info("Loading %s", TARGET_URL)
        try_goto(page, TARGET_URL)

        # Scroll down to trigger lazy loading and load pagination
        for _ in range(5):
            page.mouse.wheel(0, 3000)
            time.sleep(0.6)

        total_pages = _get_total_pages(page)
        log.info("Detected %d total pages", total_pages)

        # Collect products from page 1
        page1_products = collect_dom_listing(page, TARGET_URL)
        all_listing_products.extend(page1_products)
        log.info("Page 1: collected %d products from DOM", len(page1_products))

        # Navigate through remaining pages by CLICKING pagination buttons
        # (the site is a SPA — URL navigation doesn't change displayed products)
        global_seen_ids: Set[str] = {p.get("product_id", "") for p in page1_products}
        stale_page_count = 0
        for page_num in range(2, total_pages + 1):
            if limit > 0 and len(all_listing_products) >= limit:
                break

            log.info("Clicking to page %d/%d", page_num, total_pages)
            if not _click_next_page(page, page_num):
                log.warning("Could not click to page %d — trying URL fallback", page_num)
                # Fallback: try URL navigation
                page_url = f"{TARGET_URL}?page={page_num}"
                try:
                    try_goto(page, page_url, retries=2)
                    for _ in range(3):
                        page.mouse.wheel(0, 3000)
                        time.sleep(0.4)
                except Exception as exc:
                    log.warning("URL fallback also failed for page %d: %s", page_num, exc)
                    continue

            # Scroll to ensure all lazy content loads
            for _ in range(3):
                page.mouse.wheel(0, 2000)
                time.sleep(0.3)

            page_products = collect_dom_listing(page, TARGET_URL)
            # Filter out products we already have (dedup across pages)
            new_products = [
                p for p in page_products
                if p.get("product_id", "") not in global_seen_ids
            ]
            if new_products:
                all_listing_products.extend(new_products)
                global_seen_ids.update(p.get("product_id", "") for p in new_products)
                stale_page_count = 0
                log.info("Page %d: collected %d new products (total DOM: %d)",
                         page_num, len(new_products), len(all_listing_products))
            else:
                stale_page_count += 1
                log.info("Page %d: no new products (stale count: %d)", page_num, stale_page_count)
                if stale_page_count >= 3:
                    log.info("Stopping pagination: 3 consecutive pages with no new products")
                    break

            time.sleep(random.uniform(0.3, 0.8))

        browser.close()

    log.info(
        "DOM listing: %d products | API discovery: %d products | %d endpoints",
        len(all_listing_products), len(discovered_products), len(endpoint_urls),
    )

    # Save discovered API endpoints
    api_endpoint_list = sorted(endpoint_urls)
    api_endpoints_path.parent.mkdir(parents=True, exist_ok=True)
    api_endpoints_path.write_text(json.dumps(api_endpoint_list, indent=2), encoding="utf-8")

    # Try fetching from discovered API endpoints (with pagination)
    fetched_api_products = scrape_from_discovered_endpoints(api_endpoint_list)
    preferred = discovered_products + fetched_api_products

    normalized_rows: List[Dict[str, Any]] = []
    dedupe: Set[str] = set()

    # Prefer API-sourced products (richer data)
    if preferred:
        log.info("Using %d API-sourced products as primary source", len(preferred))
        for item in preferred:
            normalized = normalize_product(item, TARGET_URL)
            if not normalized:
                continue
            key = normalized["product_id"] or normalized["url"]
            if key in dedupe:
                continue
            dedupe.add(key)
            normalized_rows.append(normalized)
            if limit > 0 and len(normalized_rows) >= limit:
                break

    # Also add DOM-scraped products (may have additional items or fill gaps)
    log.info("Processing %d DOM-listed products", len(all_listing_products))
    for item in all_listing_products:
        normalized = normalize_product(item, TARGET_URL)
        if not normalized:
            continue
        key = normalized["product_id"] or normalized["url"]
        if key in dedupe:
            continue
        dedupe.add(key)
        normalized_rows.append(normalized)
        if limit > 0 and len(normalized_rows) >= limit:
            break

    # If we still have products that need detail enrichment and we don't have
    # enough data from API, scrape individual product pages
    products_needing_detail = [
        row for row in normalized_rows
        if not row.get("category") and not row.get("unit") and row.get("url")
    ]
    if products_needing_detail and len(products_needing_detail) <= 500:
        log.info("Enriching %d products via detail page scraping", len(products_needing_detail))
        with sync_playwright() as p:
            browser = p.chromium.launch(headless=headless, slow_mo=slowmo)
            context = browser.new_context(
                user_agent="Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                           "AppleWebKit/537.36 (KHTML, like Gecko) "
                           "Chrome/122.0.0.0 Safari/537.36"
            )
            detail_page = context.new_page()
            count = 0
            for product in products_needing_detail:
                detailed = scrape_product_detail(detail_page, product)
                if detailed:
                    # Update the existing row in-place
                    for i, row in enumerate(normalized_rows):
                        if row["product_id"] == detailed["product_id"]:
                            normalized_rows[i] = detailed
                            break
                count += 1
                if count % 20 == 0:
                    log.info("Detail scrape progress: %d/%d", count, len(products_needing_detail))
                    # Incremental save
                    merged = list(existing.values()) + normalized_rows
                    merged_dedup = {}
                    for row in merged:
                        k = str(row.get("product_id") or "") or str(row.get("url") or "")
                        if k:
                            merged_dedup[k] = row
                    write_products(output, list(merged_dedup.values()))
            browser.close()

    # Merge with existing data and write final output
    merged_map = dict(existing)
    for row in normalized_rows:
        key = row["product_id"] or row["url"]
        merged_map[key] = row

    final_rows = list(merged_map.values())
    write_products(output, final_rows)
    log.info("Saved %d total products to %s", len(final_rows), output)
    return len(final_rows)


def main() -> None:
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
    )
    parser = argparse.ArgumentParser(description="Scrape Bazarey products to XLSX")
    parser.add_argument("--headless", default="true", help="true/false")
    parser.add_argument("--limit", type=int, default=0, help="0 means no limit")
    parser.add_argument("--output", default="", help="Output XLSX path")
    parser.add_argument("--slowmo", type=int, default=0, help="Playwright slow motion in ms")
    args = parser.parse_args()

    settings = Settings.from_env()
    output = Path(args.output) if args.output else settings.products_xlsx

    total = run_scraper(
        headless=parse_bool(args.headless),
        limit=args.limit,
        output=output,
        slowmo=args.slowmo,
        api_endpoints_path=settings.api_endpoints_json,
    )
    print(f"Saved {total} products to {output}")


if __name__ == "__main__":
    main()
