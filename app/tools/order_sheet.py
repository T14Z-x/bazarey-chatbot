from __future__ import annotations

from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional

import portalocker
from openpyxl import Workbook, load_workbook

ORDER_HEADERS = [
    "order_id",
    "created_at",
    "channel",
    "channel_user_id",
    "customer_name",
    "phone",
    "address",
    "area",
    "items",
    "total",
    "notes",
    "status",
    "last_message",
]


def utc_now_iso() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat()


class OrderSheet:
    def __init__(self, path: Path) -> None:
        self.path = Path(path)
        self.lock_path = self.path.with_suffix(self.path.suffix + ".lock")

    def ensure_file(self) -> None:
        self.path.parent.mkdir(parents=True, exist_ok=True)
        if self.path.exists():
            return
        wb = Workbook()
        ws = wb.active
        ws.title = "orders"
        ws.append(ORDER_HEADERS)
        wb.save(self.path)

    def _read_rows(self) -> List[Dict[str, Any]]:
        self.ensure_file()
        wb = load_workbook(self.path)
        ws = wb.active
        headers = [c.value for c in ws[1]]
        rows: List[Dict[str, Any]] = []
        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            item = dict(zip(headers, row))
            item["_row_index"] = idx
            rows.append(item)
        wb.close()
        return rows

    def _next_order_id(self, rows: List[Dict[str, Any]]) -> str:
        max_num = 0
        for row in rows:
            order_id = str(row.get("order_id") or "")
            if order_id.startswith("BZ-"):
                try:
                    max_num = max(max_num, int(order_id.split("-")[1]))
                except Exception:
                    continue
        return f"BZ-{max_num + 1:06d}"

    def get_active_order(self, channel_user_id: str) -> Optional[Dict[str, Any]]:
        for row in reversed(self._read_rows()):
            if row.get("channel_user_id") != channel_user_id:
                continue
            status = str(row.get("status") or "").upper()
            if status not in {"CONFIRMED", "CANCELLED"}:
                return row
        return None

    def set_status(self, order_id: str, status: str, last_message: str = "") -> Optional[Dict[str, Any]]:
        self.ensure_file()
        with portalocker.Lock(self.lock_path, timeout=10):
            wb = load_workbook(self.path)
            ws = wb.active
            headers = [c.value for c in ws[1]]
            header_to_col = {name: idx + 1 for idx, name in enumerate(headers)}
            found = None
            for row_idx in range(2, ws.max_row + 1):
                if ws.cell(row=row_idx, column=header_to_col["order_id"]).value == order_id:
                    ws.cell(row=row_idx, column=header_to_col["status"]).value = status
                    ws.cell(row=row_idx, column=header_to_col["last_message"]).value = last_message
                    found = row_idx
                    break
            wb.save(self.path)
            wb.close()

        if found is None:
            return None
        for row in self._read_rows():
            if row.get("order_id") == order_id:
                return row
        return None

    def upsert_active_order(
        self,
        channel_user_id: str,
        payload: Dict[str, Any],
        status: str,
    ) -> Dict[str, Any]:
        self.ensure_file()
        order_id = ""
        with portalocker.Lock(self.lock_path, timeout=10):
            wb = load_workbook(self.path)
            ws = wb.active
            headers = [c.value for c in ws[1]]
            header_to_col = {name: idx + 1 for idx, name in enumerate(headers)}

            active_row = None
            for row_idx in range(ws.max_row, 1, -1):
                user = ws.cell(row=row_idx, column=header_to_col["channel_user_id"]).value
                row_status = str(ws.cell(row=row_idx, column=header_to_col["status"]).value or "").upper()
                if user == channel_user_id and row_status not in {"CONFIRMED", "CANCELLED"}:
                    active_row = row_idx
                    break

            if active_row is None:
                rows_for_id = []
                for row_idx in range(2, ws.max_row + 1):
                    rows_for_id.append({"order_id": ws.cell(row=row_idx, column=header_to_col["order_id"]).value})
                order_id = self._next_order_id(rows_for_id)
                active_row = ws.max_row + 1
                ws.cell(row=active_row, column=header_to_col["order_id"]).value = order_id
                ws.cell(row=active_row, column=header_to_col["created_at"]).value = utc_now_iso()
                ws.cell(row=active_row, column=header_to_col["channel"]).value = payload.get("channel", "simulator")
                ws.cell(row=active_row, column=header_to_col["channel_user_id"]).value = channel_user_id
            else:
                order_id = str(ws.cell(row=active_row, column=header_to_col["order_id"]).value or "")

            for key, value in payload.items():
                if key not in header_to_col:
                    continue
                ws.cell(row=active_row, column=header_to_col[key]).value = value

            ws.cell(row=active_row, column=header_to_col["status"]).value = status
            wb.save(self.path)
            wb.close()

        if not order_id:
            raise RuntimeError("Failed to resolve order id after upsert")

        for row in reversed(self._read_rows()):
            if str(row.get("order_id") or "") == order_id:
                return row

        raise RuntimeError("Failed to upsert order row")
