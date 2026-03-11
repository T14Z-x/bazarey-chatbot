from __future__ import annotations

import re
from pathlib import Path
from typing import Any, Dict, List, Optional

from openpyxl import Workbook, load_workbook

try:
    from rapidfuzz import fuzz
except Exception:  # pragma: no cover
    fuzz = None

try:
    from app.bot.normalizer import normalize_query as _normalize_query
except Exception:  # pragma: no cover
    def _normalize_query(t: str) -> str:  # type: ignore[misc]
        return t

from app.tools.vector_store import VectorStore

PRODUCT_HEADERS = [
    "product_id",
    "name",
    "url",
    "category",
    "price",
    "regular_price",
    "unit",
    "stock_qty",
    "is_active",
    "image_url",
    "updated_at",
]


class ProductCatalog:
    def __init__(self, path: Path, vector_index_path: Optional[Path] = None) -> None:
        self.path = Path(path)
        self._cache: List[Dict[str, Any]] = []
        self._mtime: Optional[float] = None
        self.vector_store = VectorStore(vector_index_path) if vector_index_path else None

    def ensure_file(self) -> None:
        self.path.parent.mkdir(parents=True, exist_ok=True)
        if self.path.exists():
            return
        wb = Workbook()
        ws = wb.active
        ws.title = "products"
        ws.append(PRODUCT_HEADERS)
        wb.save(self.path)

    def _normalize(self, text: str) -> str:
        text = (text or "").lower()
        text = re.sub(r"[^\w\s\u0980-\u09FF]", " ", text)
        return re.sub(r"\s+", " ", text).strip()

    def _score(self, query: str, candidate: str) -> float:
        query_n = self._normalize(query)
        cand_n = self._normalize(candidate)
        if not query_n or not cand_n:
            return 0.0
        if fuzz is not None:
            return float(fuzz.token_set_ratio(query_n, cand_n))

        q_tokens = set(query_n.split())
        c_tokens = set(cand_n.split())
        if not q_tokens or not c_tokens:
            return 0.0
        overlap = len(q_tokens & c_tokens)
        return (overlap / len(q_tokens)) * 100.0

    def _product_score(self, query: str, product: dict) -> float:
        """Score query against product name AND category for broader matching."""
        name = product.get("name", "")
        category = product.get("category", "")
        name_score = self._score(query, name)
        # Category is only a soft signal — prevents "tea" from blindly matching
        # all items in "Tea & Coffee" when the product name says "Coffee".
        cat_score = self._score(query, category) * 0.45
        return max(name_score, cat_score)

    def _coerce_float(self, value: Any) -> Optional[float]:
        if value is None or value == "":
            return None
        try:
            if isinstance(value, str):
                cleaned = re.sub(r"[^\d.,-]", "", value).replace(",", "")
                return float(cleaned)
            return float(value)
        except Exception:
            return None

    def _coerce_int(self, value: Any) -> Optional[int]:
        if value is None or value == "":
            return None
        try:
            return int(float(value))
        except Exception:
            return None

    def _load(self) -> List[Dict[str, Any]]:
        self.ensure_file()
        mtime = self.path.stat().st_mtime
        if self._cache and self._mtime == mtime:
            return self._cache

        wb = load_workbook(self.path, data_only=True)
        ws = wb.active
        rows: List[Dict[str, Any]] = []
        headers = [c.value for c in ws[1]]

        for row in ws.iter_rows(min_row=2, values_only=True):
            raw = dict(zip(headers, row))
            if not raw.get("name"):
                continue
            rows.append(
                {
                    "product_id": str(raw.get("product_id") or "").strip(),
                    "name": str(raw.get("name") or "").strip(),
                    "url": str(raw.get("url") or "").strip(),
                    "category": str(raw.get("category") or "").strip(),
                    "price": self._coerce_float(raw.get("price")),
                    "regular_price": self._coerce_float(raw.get("regular_price")),
                    "unit": str(raw.get("unit") or "").strip(),
                    "stock_qty": self._coerce_int(raw.get("stock_qty")),
                    "is_active": bool(raw.get("is_active", True)),
                    "image_url": str(raw.get("image_url") or "").strip(),
                    "updated_at": str(raw.get("updated_at") or "").strip(),
                }
            )
        wb.close()

        self._cache = rows
        self._mtime = mtime

        # Update vector index if it's missing or outdated
        if self.vector_store and not self.vector_store.load():
            self.vector_store.build_index(rows)

        return rows

    def search_products(
        self, query: str, limit: int = 5, min_score: float = 35,
    ) -> List[Dict[str, Any]]:
        """Search products with Hybrid Matching (Fuzzy + Semantic)."""
        normalized_query = _normalize_query(query)
        products = self._load()
        
        # 1. Fuzzy Search Results
        fuzzy_results: Dict[str, float] = {}
        for p in products:
            if not p.get("is_active", True):
                continue
            score = max(
                self._product_score(normalized_query, p),
                self._product_score(query, p),
            )
            if score >= min_score:
                fuzzy_results[p["product_id"]] = score

        # 2. Semantic Search Results
        vector_results: Dict[str, float] = {}
        if self.vector_store:
            v_matches = self.vector_store.search(query, top_k=limit * 2)
            for m in v_matches:
                vector_results[m["product_id"]] = m["score"]

        # 3. Combine results (Hybrid)
        combined: Dict[str, float] = {}
        all_ids = set(fuzzy_results.keys()) | set(vector_results.keys())
        
        for pid in all_ids:
            f_score = fuzzy_results.get(pid, 0)
            v_score = vector_results.get(pid, 0)
            
            # If item is in both, boost it. Otherwise take the max or weighted sum.
            if f_score > 0 and v_score > 0:
                combined[pid] = max(f_score, v_score) * 1.2
            else:
                combined[pid] = max(f_score, v_score)

        # 4. Final Ranking
        ranked = []
        for pid, score in combined.items():
            product = self.get_product(pid)
            if product:
                ranked.append((score, product))

        ranked.sort(key=lambda x: x[0], reverse=True)
        
        results = []
        for score, product in ranked[:limit]:
            results.append(
                {
                    "product_id": product["product_id"],
                    "name": product["name"],
                    "price": product["price"],
                    "unit": product["unit"],
                    "stock_qty": product["stock_qty"],
                    "url": product["url"],
                    "_score": round(score, 1),
                }
            )
        return results

    def list_categories(self) -> List[Dict[str, Any]]:
        """Return all categories with product counts, sorted by count desc."""
        products = self._load()
        counts: Dict[str, int] = {}
        for p in products:
            if not p.get("is_active", True):
                continue
            cat = (p.get("category") or "").strip()
            if cat:
                counts[cat] = counts.get(cat, 0) + 1
        result = [{"category": k, "count": v} for k, v in counts.items()]
        result.sort(key=lambda x: x["count"], reverse=True)
        return result

    def browse_category(
        self, category_query: str, limit: int = 10,
    ) -> Dict[str, Any]:
        """Find a category by fuzzy match and return its products."""
        # Normalise the query
        normalized = _normalize_query(category_query)
        products = self._load()

        # Collect unique categories
        cat_set: Dict[str, str] = {}  # normalized_lower -> original
        for p in products:
            cat = (p.get("category") or "").strip()
            if cat and cat not in cat_set.values():
                cat_set[self._normalize(cat)] = cat

        # Fuzzy-match the query against category names
        best_cat = ""
        best_score = 0.0
        for cat_norm, cat_orig in cat_set.items():
            s1 = self._score(normalized, cat_orig)
            s2 = self._score(category_query, cat_orig)
            s = max(s1, s2)
            if s > best_score:
                best_score = s
                best_cat = cat_orig

        if best_score < 35 or not best_cat:
            return {"category": None, "products": [], "message": f"No matching category for '{category_query}'."}

        # Get products in that category
        items = []
        for p in products:
            if p.get("category") == best_cat and p.get("is_active", True):
                items.append({
                    "product_id": p["product_id"],
                    "name": p["name"],
                    "price": p["price"],
                    "unit": p["unit"],
                    "stock_qty": p["stock_qty"],
                    "url": p.get("url", ""),
                })
        items.sort(key=lambda x: x["name"])
        selected = items if limit <= 0 else items[:limit]
        return {
            "category": best_cat,
            "count": len(items),
            "products": selected,
        }

    def get_product(self, product_id: str) -> Optional[Dict[str, Any]]:
        for product in self._load():
            if product.get("product_id") == product_id:
                return product
        return None

    def quote_items(self, items: List[Dict[str, Any]]) -> Dict[str, Any]:
        lines = []
        subtotal = 0.0
        for item in items:
            product_id = str(item.get("product_id") or "")
            qty = int(item.get("qty") or 0)
            if qty <= 0:
                continue
            product = self.get_product(product_id)
            if not product or product.get("price") is None:
                continue
            line_total = float(product["price"]) * qty
            subtotal += line_total
            lines.append(
                {
                    "product_id": product_id,
                    "name": product["name"],
                    "qty": qty,
                    "unit_price": float(product["price"]),
                    "line_total": line_total,
                    "unit": product.get("unit") or "",
                }
            )
        return {"line_items": lines, "subtotal": round(subtotal, 2)}
