from __future__ import annotations

from pathlib import Path
from typing import Any, Dict, List

from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook

from app.config import Settings
from app.main import create_app


class FakeOllamaClient:
    """Deterministic local fake used for tests."""

    def chat_json(self, messages: List[Dict[str, str]], max_retries: int = 2) -> Dict[str, Any]:
        latest = messages[-1]["content"].lower()

        if '"tool": "search_products"' in latest:
            return {
                "type": "final",
                "message": "Checked catalog. Please share quantity if you want to order.",
            }

        if "price" in latest or "দাম" in latest:
            return {"type": "tool_call", "tool": "search_products", "args": {"query": "miniket rice 5kg", "limit": 3}}

        return {"type": "final", "message": "Noted."}


def build_products(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "products"
    ws.append(
        [
            "product_id",
            "name",
            "url",
            "category",
            "price",
            "regular_price",
            "unit",
            "stock_qty",
            "is_active",
            "image_url",
            "updated_at",
        ]
    )
    ws.append([
        "p-001",
        "Miniket Rice 5kg",
        "https://example.com/miniket-rice-5kg",
        "Rice",
        450,
        "",
        "5kg",
        30,
        True,
        "",
        "2026-01-01T00:00:00+00:00",
    ])
    wb.save(path)


def test_simulated_order_flow(tmp_path: Path) -> None:
    data_dir = tmp_path / "data"
    data_dir.mkdir(parents=True, exist_ok=True)

    settings = Settings(
        base_dir=tmp_path,
        data_dir=data_dir,
        products_xlsx=data_dir / "products.xlsx",
        orders_xlsx=data_dir / "orders.xlsx",
        sessions_db=data_dir / "sessions.db",
        api_endpoints_json=data_dir / "api_endpoints.json",
        llm_provider="ollama",
        groq_api_key="",
        groq_model="llama-3.1-8b-instant",
        ollama_host="http://localhost:11434",
        ollama_model="llama3.1:8b",
    )

    build_products(settings.products_xlsx)

    app = create_app(settings=settings, llm_client=FakeOllamaClient())
    client = TestClient(app)

    ui = client.get("/")
    assert ui.status_code == 200
    assert "Bazarey Local Console" in ui.text

    r1 = client.post("/simulate/chat", json={"channel_user_id": "u123", "text": "miniket rice 5kg price?"})
    assert r1.status_code == 200
    assert "Miniket Rice 5kg" in r1.json()["reply"]
    assert "450" in r1.json()["reply"]

    r2 = client.post("/simulate/chat", json={"channel_user_id": "u123", "text": "2 ta nibo"})
    assert r2.status_code == 200
    assert "কার্টে যোগ হয়েছে" in r2.json()["reply"]

    r3 = client.post("/simulate/chat", json={"channel_user_id": "u123", "text": "order"})
    assert r3.status_code == 200
    assert "নাম লিখুন" in r3.json()["reply"]

    r4 = client.post("/simulate/chat", json={"channel_user_id": "u123", "text": "Rahim"})
    assert r4.status_code == 200
    assert "ফোন" in r4.json()["reply"]

    r5 = client.post("/simulate/chat", json={"channel_user_id": "u123", "text": "01700111222"})
    assert r5.status_code == 200
    assert "ঠিকানা" in r5.json()["reply"]

    r6 = client.post("/simulate/chat", json={"channel_user_id": "u123", "text": "House 10, Road 5, Dhanmondi, Dhaka"})
    assert r6.status_code == 200
    assert "YES" in r6.json()["reply"]

    r7 = client.post("/simulate/chat", json={"channel_user_id": "u123", "text": "YES"})
    assert r7.status_code == 200
    assert "কনফার্ম" in r7.json()["reply"]

    wb = load_workbook(settings.orders_xlsx, data_only=True)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    rows = [dict(zip(headers, row)) for row in ws.iter_rows(min_row=2, values_only=True)]
    wb.close()

    assert rows, "orders.xlsx should contain at least one row"
    last = rows[-1]
    assert last["channel_user_id"] == "u123"
    assert last["status"] == "CONFIRMED"
    assert float(last["total"]) == 900.0
    assert "Miniket Rice 5kg" in str(last["items"])


def test_banglish_followup_understanding(tmp_path: Path) -> None:
    data_dir = tmp_path / "data"
    data_dir.mkdir(parents=True, exist_ok=True)

    settings = Settings(
        base_dir=tmp_path,
        data_dir=data_dir,
        products_xlsx=data_dir / "products.xlsx",
        orders_xlsx=data_dir / "orders.xlsx",
        sessions_db=data_dir / "sessions.db",
        api_endpoints_json=data_dir / "api_endpoints.json",
        llm_provider="ollama",
        groq_api_key="",
        groq_model="llama-3.1-8b-instant",
        ollama_host="http://localhost:11434",
        ollama_model="llama3.1:8b",
    )

    build_products(settings.products_xlsx)

    app = create_app(settings=settings, llm_client=FakeOllamaClient())
    client = TestClient(app)

    # Product mention should return the "নিতে চাইলে বলুন" style prompt.
    r1 = client.post("/simulate/chat", json={"channel_user_id": "u-bn", "text": "miniket rice 5kg"})
    assert r1.status_code == 200
    assert "Miniket Rice 5kg" in r1.json()["reply"]

    # Banglish take-intent should ask for quantity.
    r2 = client.post("/simulate/chat", json={"channel_user_id": "u-bn", "text": "nite chai"})
    assert r2.status_code == 200
    assert "কতটি" in r2.json()["reply"]

    # Bare Bangla-digit quantity should be accepted while awaiting qty.
    r3 = client.post("/simulate/chat", json={"channel_user_id": "u-bn", "text": "৫"})
    assert r3.status_code == 200
    assert "কার্টে যোগ হয়েছে" in r3.json()["reply"]

    r4 = client.post("/simulate/chat", json={"channel_user_id": "u-bn", "text": "show cart"})
    assert r4.status_code == 200
    assert "Miniket Rice 5kg" in r4.json()["reply"]
    assert "x5" in r4.json()["reply"]
